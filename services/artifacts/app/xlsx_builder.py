"""SheetContent -> .xlsx bytes (openpyxl). Keeps a live SUM formula, not just values."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import brand
from .models import SheetContent

THIN = Side(style="thin", color=brand.BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _neutralize(v):
    """Prevent CSV/formula injection: a string cell starting with = + - @ would
    become a live formula in Excel. Prefix with ' so Excel treats it as text.
    (Our own trusted =SUM is written separately and is not passed through here.)"""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def build_sheet(content: SheetContent, name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (content.title or "Sheet")[:31]

    ncols = len(content.columns)

    # title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=content.title)
    tcell.font = Font(name=brand.UI_FONT, size=14, bold=True, color=brand.INK)
    tcell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    # header row (row 2)
    for c, col in enumerate(content.columns, start=1):
        cell = ws.cell(row=2, column=c, value=col)
        cell.font = Font(name=brand.UI_FONT, bold=True, color=brand.WHITE)
        cell.fill = PatternFill("solid", fgColor=brand.INDIGO)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="left" if c == 1 else "right")

    # data rows
    first_data = 3
    row_is_total: dict[int, bool] = {}
    for r, row in enumerate(content.rows, start=first_data):
        is_total = bool(row) and str(row[0]).strip().lower() == "total"
        row_is_total[r] = is_total
        for c, val in enumerate(row[:ncols], start=1):
            cell = ws.cell(row=r, column=c, value=_neutralize(val))
            cell.border = BOX
            cell.font = Font(name=brand.UI_FONT, bold=is_total or c == 1)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right")
    last_data = first_data + len(content.rows) - 1

    # Live SUM over the last numeric column. List the exact non-Total numeric
    # cells so the formula is correct regardless of where a Total row sits.
    last_numeric_col = None
    for c in range(ncols, 0, -1):
        if any(_is_number(row[c - 1]) for row in content.rows if len(row) >= c):
            last_numeric_col = c
            break
    if last_numeric_col:
        col_letter = get_column_letter(last_numeric_col)
        cells = [
            f"{col_letter}{r}"
            for r, row in zip(range(first_data, last_data + 1), content.rows)
            if not row_is_total[r] and len(row) >= last_numeric_col and _is_number(row[last_numeric_col - 1])
        ]
        if cells:
            frow = last_data + 1
            lbl = ws.cell(row=frow, column=1, value="Σ formula")
            lbl.font = Font(name=brand.UI_FONT, italic=True, color=brand.MUTED)
            fcell = ws.cell(row=frow, column=last_numeric_col, value=f"=SUM({','.join(cells)})")
            fcell.font = Font(name=brand.UI_FONT, bold=True, color=brand.INDIGO)
            fcell.alignment = Alignment(horizontal="right")

    # column widths + freeze
    ws.column_dimensions["A"].width = 22
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
